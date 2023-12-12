from pprint import pprint
from Main import text
from openpyxl import *
import math

fn = 'table.xlsx'
wb = load_workbook(fn)
ws = wb['Лист3']
symbols_to_encode = '0123456789абвгдеёжзийклмнопрстуфхцчшщъыьэюя.,:;- ('

class node:
    def __init__(self) -> None:
        self.sym = ''
        self.pro = 0.0
        self.arr = [0] * 20
        self.top = 0


p = [node() for _ in range(50)]

def calculate_entropy(probabilities):
    entropy = -sum(prob * math.log2(prob) for prob in probabilities.values())
    return entropy

def shannon_fano_average_length(codes, probabilities):
    avg_length = sum(len(codes[char]) * probabilities[char] for char in codes)
    return avg_length
def is_prefix(code, other_code):
    return code.startswith(other_code) or other_code.startswith(code)

def shannon_fano_decode(encoded_text, codes):
    reverse_codes = {code: symbol for symbol, code in codes.items()}

    current_code = ""
    decoded_text = ""

    for bit in encoded_text:
        current_code += bit
        if current_code in reverse_codes:
            decoded_text += reverse_codes[current_code]
            current_code = ""
    print(decoded_text)
    return decoded_text

def shannon_fano_encode(text, symbols_to_encode):
    probabilities = final
    encoded_text = ''.join(probabilities[symbol] for symbol in text if symbol in symbols_to_encode)

    return encoded_text, probabilities


def check_decoding_uniqueness(codes):
    for code1, symbol1 in codes.items():
        for code2, symbol2 in codes.items():
            if code1 != code2 and (code1.startswith(code2) or code2.startswith(code1)):
                return False
    return True

def calculate_probabilities(text, symbols_to_encode):
    row = 2
    filtered_text = ''.join(char for char in text if char in symbols_to_encode)

    symbol_frequencies = [(symbol, filtered_text.count(symbol)) for symbol in set(filtered_text)]
    symbol_frequencies = sorted(symbol_frequencies, key=lambda x: x[1], reverse=True)

    total_characters = len(filtered_text)
    probabilities = {}
    for symbol, frequency in symbol_frequencies:
        probabilities[symbol] = frequency/total_characters

    for key,value in probabilities.items():
        value_1 = key
        value_2 = value
        cell = ws.cell(row=row,column=1)
        cell.value = value_1
        cell_2 = ws.cell(row=row,column=2)
        cell_2.value = value_2
        row+=1

    return probabilities


def text_symb_calc(text, symbols_to_encode):
    count_symb = 0
    text_sort = set(text)
    for i in symbols_to_encode:
        if i in text_sort:
            count_symb += 1
    return count_symb
def shannon(l, h, p):
    pack1 = 0;
    pack2 = 0;
    diff1 = 0;
    diff2 = 0
    if ((l + 1) == h or l == h or l > h):
        if (l == h or l > h):
            return
        p[h].top += 1
        p[h].arr[(p[h].top)] = 0
        p[l].top += 1
        p[l].arr[(p[l].top)] = 1

        return

    else:
        for i in range(l, h):
            pack1 = pack1 + p[i].pro
        pack2 = pack2 + p[h].pro
        diff1 = pack1 - pack2
        if (diff1 < 0):
            diff1 = diff1 * -1
        j = 2
        while (j != h - l + 1):
            k = h - j
            pack1 = pack2 = 0
            for i in range(l, k + 1):
                pack1 = pack1 + p[i].pro
            for i in range(h, k, -1):
                pack2 = pack2 + p[i].pro
            diff2 = pack1 - pack2
            if (diff2 < 0):
                diff2 = diff2 * -1
            if (diff2 >= diff1):
                break
            diff1 = diff2
            j += 1

        k += 1
        for i in range(l, k + 1):
            p[i].top += 1
            p[i].arr[(p[i].top)] = 1

        for i in range(k + 1, h + 1):
            p[i].top += 1
            p[i].arr[(p[i].top)] = 0


        shannon(l, k, p)
        shannon(k + 1, h, p)



def sortByProbability(n, p):
    temp = node()
    for j in range(1, n):
        for i in range(n - 1):
            if ((p[i].pro) > (p[i + 1].pro)):
                temp.pro = p[i].pro
                temp.sym = p[i].sym

                p[i].pro = p[i + 1].pro
                p[i].sym = p[i + 1].sym

                p[i + 1].pro = temp.pro
                p[i + 1].sym = temp.sym



def final_calc(n):
    row = 2
    fin = {}
    for i in range(n - 1, -1, -1):
        code = ''
        for j in range(p[i].top + 1):
            code += str(p[i].arr[j])
        fin[p[i].sym] = code
    for i in fin.values():
        value = i
        cell = ws.cell(row=row,column=3)
        cell.value = value
        row += 1
    return(fin)



total = 0
probabilities = calculate_probabilities(text,symbols_to_encode)

n = text_symb_calc(text,symbols_to_encode)
i = 0
for symb in probabilities.keys():
    p[i].sym += symb
    i += 1
x = []

for k in probabilities.keys():
    x.append(probabilities[k])
for i in range(n):
    p[i].pro = x[i]
    total = total + p[i].pro


    if (total > 1):
        total = total - p[i].pro
        i -= 1

i += 1
p[i].pro = 1 - total

sortByProbability(n, p)

for i in range(n):
    p[i].top = -1
final = final_calc(n)
enc = shannon_fano_encode(text,symbols_to_encode)
print(enc)
print(shannon_fano_decode(enc,final))
shannon(0, n - 1, p)
calculate_probabilities(text, symbols_to_encode)

wb.save(fn)
wb.close()
